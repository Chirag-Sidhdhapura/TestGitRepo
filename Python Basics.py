"""a, b, c = 5, 4.8, "Helloe"

#print(str(a), str(b), c, sep='-')

#print("Abcd", "xyz", sep= "Seperator" )

print("{} - {}".format("Chirag",b))

print(type(b))

name = "Alice"
age = 30
print("Name:", name, "Age:", age)

abc= 1001 +  5
print(type(abc))


L = [1, 3, "Hello", 4.5, 'c']
del L[1:4]
#L.append("Appended")
#print(L.index(4.5))
#L.insert(1, 555)
print(L)

s = "Chirag Sidhdhapura"
l = list(set(list(s)))
l.sort()

print(''.join(l).replace(" ", "").lower())
"""
import openpyxl
"""
l = "Chirag sidhdhapura best SDET out there in the market".split(" ")
print(l)
print(l.reverse())
print(l)

abcd = [4, 5, 6, 2, 55, 4, 1, 0]
print(abcd)
print(tuple(abcd))
print(set(abcd)) """
dict = {}
workbook = openpyxl.load_workbook(r"C:\Users\Chirag M. Sidhdhapur\Desktop\Selenium Python\Python Learning\FrameworkDesignPractise\TestData\testdata.xlsx")
sheet = workbook.active
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "OrderProduct":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict)
print("This is the changes intrtroduced by User Y")
print("This will not be visible for User X right now")
print("after pulling updates User X can see this")
print("once you can see this changes let me know by adding print statement @userX")